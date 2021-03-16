import openpyxl
from django.db.models import Q
from django.shortcuts import render,redirect
from openpyxl import load_workbook
from io import BytesIO
from User_app.models import Files_Data
from User_app.forms import Excel_Fileform
from django.contrib import messages
from django.http import JsonResponse

def index(request):
    All_Data = Files_Data.objects.all()[:10]
    context = {'Data': All_Data}
    # Files_Data.objects.all().delete()
    return render(request, 'User/index.html',context)


def Ajax_FileUpload(request):
    print('re')
    if request.is_ajax():
        excel_file = request.FILES["excel_file"]
        if excel_file:
            print(excel_file,'file')
            if str(excel_file).endswith(".xlsx"):
                wb = openpyxl.load_workbook(excel_file)
                sheets = wb.sheetnames
                worksheet = sheets
                for worksheets in wb.sheetnames:
                    worksheet = wb[worksheets]
                excel_data = list()
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                        # print(cell.value)
                    excel_data.append(row_data)
                excel_data.pop(0)
                for i in excel_data:
                    pass
                    # print(i[0],'Name')
                    # print(i[1],'Ifsc')
                    # print(i[2],'Micr')
                    # print(i[3],'Branch')
                    # print(i[4],'Address')
                    # print(i[5],'City1')
                    # print(i[6],'City2')
                    # print(i[7],'Std')
                    # print(i[8],'Phone')
                    Files_Data(BANK=i[0],IFSC=i[1],MICR=i[2],BRANCH=i[3],ADDRESS=i[4],CITY1=i[5],CITY2=i[6],STATE=[7],STD=i[8],PHONE=i[9]).save()
                mess = {'successs':'File Upload Successfully'}
            else:
                mess = {'errorss':'File Does Not Support'}
        else:
            mess = {'errorss': 'PLease Select File'}
    else:
        mess = {'errorss': 'Something Error'}
    return JsonResponse(mess)


def Search(request):
    Search_Data = request.GET.get('search')
    All_Data = Files_Data.objects.all()[:10]
    Searched_Data = Files_Data.objects.filter(Q(BANK__icontains=Search_Data)|Q(IFSC__icontains=Search_Data))
    if Searched_Data:
        context = {'Data': All_Data,'Searched_Data':Searched_Data,'Searched_input':Search_Data}
        return render(request, 'User/index.html', context)
    else:
        context = {'Data': All_Data,'Searched_input':Search_Data}
        return render(request, 'User/index.html', context)
